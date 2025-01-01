from openpyxl import Workbook

# Criando arquivo Excel
wb = Workbook()
print(wb.sheetnames)  

# Selecionando uma aba
sheet = wb["Sheet"]

# inserindo dados pelo rótulo da celula
sheet["A1"].value = "Python"
sheet["B1"].value = "Formação Expert"
sheet["C1"].value = "Automação de processos"

# inserir dados peloa coordenada da celula
sheet.cell(row=2, column=1).value = "Empowerdata"
sheet.cell(2,2).value ="Empowerdatapython"

#apagando o dado de uma celula
sheet["A1"].value = None
sheet.cell(1,3).value = None

# Salvando o arquivo
wb.save("comando_basico.xlsx")
print("Arquivo salvo com sucesso!")
