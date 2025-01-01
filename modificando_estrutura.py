from openpyxl import load_workbook

#carregando o arquivo Excel
wb = load_workbook("Modificando estrutura.xlsx")

# Verificando as abas
print(wb.sheetnames)

# selecionando a aba de vendas
sheet = wb["Vendas 2024"]

# renomenado uma aba
sheet.title = "Vendas 2024"

# criando uma nova aba
wb.create_sheet("Vendas 2025")

# mesclando celulas
sheet.merge_cells("A1:D1")

# tirando a mesclagem de celulas
sheet.unmerge_cells("A1:D1")

# inserindo linhas
sheet.insert_rows(3)
sheet.insert_rows(14)

#inserindo colunas
sheet.insert_cols(2)

#deletando linhas e ciolunas
sheet.delete_rows(3)
sheet.delete_cols(2)

wb.save("Modificando estrutura.xlsx")